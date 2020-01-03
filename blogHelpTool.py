#Import libraries
import os
import sys
import bs4
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import workbook

#Initialize variables
filename = ''
linkHeaders = ''
failClause = ''

columnCount = 0

argumentCount = len(sys.argv)

artOperant = True
gameOperant = True
playerOperant = True
timeOperant = True
ratingOperant = True

columnEmpty = True

#.xlsx argv check
if argumentCount > 1:
    for n in (1, argumentCount - 1):
        if str(sys.argv[n])[-5:] == '.xlsx':
            filename = sys.argv[n]
        if str(sys.argv[n]) == '-a':
            artOperant = False
        if str(sys.argv[n]) == '-g':
            gameOperant = False
        if str(sys.argv[n]) == '-p':
            playerOperant = False
        if str(sys.argv[n]) == '-t':
            timeOperant = False
        if str(sys.argv[n]) == '-r':
            ratingOperant = False

while True:#Loop input until input leads to a valid file.

    print('\nPlease input the path to a valid .xlsx file.\n\t(Valid .xlsx files specifically have 5 columns of data.)')
    filename = input()
    print('Validating file...')

    if filename[-5:] == '.xlsx':
        #Since the file given was an xlsx file, create variables based on the spreadsheet given.
        book = openpyxl.load_workbook(filename)
        sheet = book.active
<<<<<<< HEAD
=======
<<<<<<< HEAD
>>>>>>> 95a110c5272db22c4df1a4a5a6ad473f00afedd3
        width = sheet.max_column
        height = sheet.max_row

        width = sheet.max_column
        height = sheet.max_row

        for x in range(1, width + 1):
            for y in range (1, height + 1):
                if sheet.cell(row=y, column=x).value != 0:
                    columnEmpty = False

            if not columnEmpty:
                columnCount += 1

        if columnCount == 5:
            print('File validated!\n\tContinuing...')
            break

        else:
            failClause = 'The file given had ' + str(columnCount) + ' non-empty columns. Please change this to 5 non-empty columns.'

    else:
        failClause = 'User did not give an .xlsx file.'

    #Failure case below. Print out the reason.
    print('Process failed. Reason:\n\t%s' % failClause)

#Once the file has been validated, the code continues to process the sheet.
<<<<<<< HEAD
=======
=======
>>>>>>> 3e2da730007347a698a24135bb0651e44c70309d

        width = sheet.max_column
        height = sheet.max_row

        for x in range(1, width + 1):
            for y in range (1, height + 1):
                if sheet.cell(row=y, column=x).value != 0:
                    columnEmpty = False

            if not columnEmpty:
                columnCount += 1

        if columnCount == 5:
            print('File validated!\n\tContinuing...')
            break

        else:
            failClause = 'The file given had ' + str(columnCount) + ' non-empty columns. Please change this to 5 non-empty columns.'

    else:
        failClause = 'User did not give an .xlsx file.'

    #Failure case below. Print out the reason.
    print('Process failed. Reason:\n\t%s' % failClause)

#Once the file has been validated, the code continues to process the sheet.
>>>>>>> 95a110c5272db22c4df1a4a5a6ad473f00afedd3

#Store data on spreadsheetText variable
spreadsheetText = '[su_table class="custom-su-table" responsive="yes"]\n<table>\n<tr class="header'

if artOperant:
    spreadsheetText += '\n<td>Box Art</td>'
if gameOperant:
    spreadsheetText += '\n<td>Game</td>'
if playerOperant:
    spreadsheetText += '\n<td>Player(s)</td>'
if timeOperant:
    spreadsheetText += '\n<td>Playtime</td>'
if ratingOperant :
    spreadsheetText += '\n<td>Rating</td>'
spreadsheetText += '\n</tr>'


    #Height variable
print(str(height - 1) + ' items found. Processing...')

for i in range(2, height+1):

    #Take a tag from column 1's data, add it to an a tag with column 2's info
    itemHtml = str(sheet.cell(row=i, column=1).value)
    soup = BeautifulSoup(itemHtml, 'html.parser')
    gameaTag = soup.a#Select first a tag in html
    for img in soup('img'):
        img.decompose()
    gameaTag.insert(1, str(sheet.cell(row=i, column=2).value))
    gameaTag = str(gameaTag)

    #Find amazonURL from a tag's href
    for a in soup('a'):
        amazonURL = a['href']


    #Write values from each column
    itemDetails = '\n\n<tr>'
    for j in range(1, 6):
        cellInfo = '\n<td>' + str(sheet.cell(row=i, column=j).value) + '</td>'

        if j == 2:  #Special clause for 2nd column's data
            cellInfo = '\n<td>' + gameaTag + '</td>'
            linkHeaders += '<h2>' + gameaTag  + '</h2>' + '\n[su_button url="' + amazonURL + '" target="blank" size="4" style="default" background="#EEC562" color="#000000" class="check-amazon"]Check availability on Amazon![/su_button]' + '\n<div style="height:100px" aria-hidden="true" class="wp-block-spacer">\n</div>\n\n'

        #Write completed cellInfo to itemDetails
        itemDetails += cellInfo

    spreadsheetText += itemDetails + '\n</tr>'

spreadsheetText += '\n</table>\n[/su_table]\n\n'
spreadsheetText += linkHeaders

#Write out value
file = open('htmlOutput.txt', 'w+')
file.write(spreadsheetText)
file.close()
