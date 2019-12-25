import os
import sys
import openpyxl
from openpyxl import workbook

content = ''

#.xlsx argv check
if sys.argv[1][-5:] != '.xlsx' or len(sys.argv) != 2:
    print("Apologies. The argument provided did not lead to a valid .xlsx file. Please try again.")
    sys.exit()
#os.exists check

#Create variables based on spreadsheet

#Store data on content variable
content += '''[su_table class="custom-su-table" responsive="yes"]
<table>
<tr class="header">
<td>Box Art</td>
<td>Game</td>
<td>Player(s)</td>
<td>Playtime</td>
<td>Rating</td>
</tr>
'''

for i in range(1, height):
    #Take href value from column 1, add it to an a tag with column 2's info
    #nameString = ^^^

    #Write values from each column
    content += '''

    <tr>
    <td>''' + sheet.cell(row=i, column=1).value + '''</td>
    <td>''' + nameString + '''</td>
    <td>''' + sheet.cell(row=i, column=3).value + '''</td>
    <td>''' + sheet.cell(row=i, column=4).value + '''</td>
    <td>''' + sheet.cell(row=i, column=5).value + '''</td>
    </tr>'''
#End the su_table tag
content += '\n[/su_table]'

#Write out value
file = open('blogHelp.txt', 'w+')
file.write(content)
file.close()

#Save
